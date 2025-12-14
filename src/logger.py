"""
📝 JOB APPLICATION LOGGER
Tracks all applications in an Excel file with statistics and analytics.
This is the bot's "memory" - helps you track what worked and what didn't.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, List
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class JobLogger:
    def __init__(self, excel_path: str = "data/job_tracker.xlsx"):
        """
        Initialize the job application logger.
        
        Args:
            excel_path: Path to Excel file for tracking applications
        """
        self.excel_path = Path(excel_path)
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Load or create workbook
        if self.excel_path.exists():
            self.workbook = openpyxl.load_workbook(self.excel_path)
            print(f"📊 Loaded existing tracker: {self.excel_path}")
        else:
            self.workbook = Workbook()
            self._initialize_sheets()
            print(f"📊 Created new tracker: {self.excel_path}")
    
    
    def _initialize_sheets(self):
        """
        Create the initial structure with multiple sheets.
        """
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        # Create Applications sheet
        apps_sheet = self.workbook.create_sheet("Applications", 0)
        self._setup_applications_sheet(apps_sheet)
        
        # Create Statistics sheet
        stats_sheet = self.workbook.create_sheet("Statistics", 1)
        self._setup_statistics_sheet(stats_sheet)
        
        # Create Daily Summary sheet
        daily_sheet = self.workbook.create_sheet("Daily Summary", 2)
        self._setup_daily_summary_sheet(daily_sheet)
        
        # Save initial structure
        self.workbook.save(self.excel_path)
    
    
    def _setup_applications_sheet(self, sheet):
        """
        Set up the Applications sheet with headers and formatting.
        """
        headers = [
            "Date", "Job Title", "Company", "Location", 
            "Resume Used", "Status", "Application URL", "Result Details"
        ]
        
        # Write headers
        sheet.append(headers)
        
        # Format headers
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = {
            'A': 18,  # Date (with time)
            'B': 35,  # Job Title
            'C': 25,  # Company
            'D': 25,  # Location
            'E': 25,  # Resume Used
            'F': 15,  # Status
            'G': 50,  # Application URL
            'H': 40,  # Result Details
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
    
    
    def _setup_statistics_sheet(self, sheet):
        """
        Set up the Statistics sheet with summary metrics.
        """
        # Title
        sheet['A1'] = "📊 JOB APPLICATION SUMMARY"
        sheet['A1'].font = Font(bold=True, size=16, color="4472C4")
        sheet.merge_cells('A1:B1')
        
        # Metrics structure
        metrics = [
            ("", ""),
            ("📝 OVERALL STATS", ""),
            ("Total Applications", "=COUNTA(Applications!A:A)-1"),
            ("Successful Submissions", "=COUNTIF(Applications!F:F,\"Success\")"),
            ("Failed Applications", "=COUNTIFS(Applications!F:F,\"*Failed*\")"),
            ("Skipped Jobs", "=COUNTIF(Applications!F:F,\"Skipped\")"),
            ("Success Rate", "=IF(B4>0,B5/B4,0)"),
            ("", ""),
            ("📁 RESUMES USED", ""),
            # Resume stats will show which resume was most successful
        ]
        
        for row_num, (label, formula) in enumerate(metrics, 1):
            sheet[f'A{row_num}'] = label
            if formula:
                sheet[f'B{row_num}'] = formula
        
        # Format section headers
        sheet['A3'].font = Font(bold=True, size=12)
        sheet['A10'].font = Font(bold=True, size=12)
        
        # Format success rate as percentage
        sheet['B8'].number_format = '0.0%'
        
        # Column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
    
    
    def _setup_daily_summary_sheet(self, sheet):
        """
        Set up the Daily Summary sheet for tracking daily progress.
        """
        headers = ["Date", "Applications", "Success", "Failed", "Success Rate"]
        
        sheet.append(headers)
        
        # Format headers
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            sheet.column_dimensions[col].width = 15
    
    
    def log_application(
        self,
        job_title: str,
        company: str,
        platform: str,
        resume_used: str,
        match_score: int,
        confidence: float,
        status: str = "Success",
        location: str = "Not specified",
        application_url: str = "",
        notes: str = ""
    ):
        """
        Log a job application with clean, user-friendly format.
        
        Args:
            job_title: Title of the job
            company: Company name
            platform: Job board (LinkedIn, Indeed, etc.) - not displayed but used for stats
            resume_used: Which resume was used (filename only)
            match_score: AI match score (0-100) - used internally but not displayed
            confidence: AI confidence (0.0-1.0) - used internally but not displayed
            status: Success/Failed/Skipped
            location: Job location
            application_url: Link to the job posting
            notes: Error messages or important details only
        """
        sheet = self.workbook["Applications"]
        
        # Prepare data - cleaner format
        now = datetime.now()
        
        # Create meaningful result details
        result_details = ""
        if status == "Success":
            result_details = "✓ Application submitted successfully"
        elif "Failed" in status:
            result_details = f"✗ {notes}" if notes else "✗ Application failed"
        elif status == "Skipped":
            result_details = f"○ Skipped - {notes}" if notes else "○ Skipped (not Easy Apply)"
        else:
            result_details = notes
        
        row_data = [
            now.strftime("%Y-%m-%d %H:%M"),     # Date with time
            job_title,                           # Job Title
            company,                             # Company
            location,                            # Location
            resume_used.replace('.pdf', ''),    # Resume (clean name)
            status,                              # Status
            application_url,                     # URL
            result_details                       # Result Details
        ]
        
        # Append to sheet
        sheet.append(row_data)
        
        # Color code status
        status_row = sheet.max_row
        status_cell = sheet.cell(row=status_row, column=6)  # Column F (Status)
        
        if status == "Success":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100", bold=True)
        elif "Failed" in status:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006", bold=True)
        else:  # Skipped
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500")
        
        # Save workbook
        self.workbook.save(self.excel_path)
        
        print(f"✅ Logged: {job_title} at {company} ({status})")
    
    
    def update_daily_summary(self):
        """
        Update the daily summary with today's statistics.
        """
        apps_sheet = self.workbook["Applications"]
        daily_sheet = self.workbook["Daily Summary"]
        
        today = datetime.now().strftime("%Y-%m-%d")
        
        # Count today's applications
        total = 0
        success = 0
        failed = 0
        
        for row in apps_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == today:  # Check date column
                total += 1
                if row[9] == "Success":  # Status column
                    success += 1
                elif row[9] == "Failed":
                    failed += 1
        
        success_rate = (success / total * 100) if total > 0 else 0
        
        # Check if today already has an entry
        existing_row = None
        for row_num, row in enumerate(daily_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == today:
                existing_row = row_num
                break
        
        # Update or append
        if existing_row:
            daily_sheet[f'B{existing_row}'] = total
            daily_sheet[f'C{existing_row}'] = success
            daily_sheet[f'D{existing_row}'] = failed
            daily_sheet[f'E{existing_row}'] = f"{success_rate:.1f}%"
        else:
            daily_sheet.append([today, total, success, failed, f"{success_rate:.1f}%"])
        
        self.workbook.save(self.excel_path)
    
    
    def get_today_count(self, platform: Optional[str] = None) -> int:
        """
        Get number of applications made today.
        
        Args:
            platform: Optional - count only for specific platform
            
        Returns:
            Number of applications today
        """
        sheet = self.workbook["Applications"]
        today = datetime.now().strftime("%Y-%m-%d")
        
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == today:  # Date column
                if platform is None or row[4] == platform:  # Platform column
                    count += 1
        
        return count
    
    
    def get_statistics(self) -> Dict[str, any]:
        """
        Get summary statistics about all applications.
        
        Returns:
            Dictionary with various statistics
        """
        sheet = self.workbook["Applications"]
        
        total = 0
        success = 0
        failed = 0
        by_platform = {}
        by_resume = {}
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            total += 1
            
            status = row[9]  # Status column
            if status == "Success":
                success += 1
            elif status == "Failed":
                failed += 1
            
            # Count by platform
            platform = row[4]
            by_platform[platform] = by_platform.get(platform, 0) + 1
            
            # Count by resume
            resume = row[6]
            by_resume[resume] = by_resume.get(resume, 0) + 1
        
        success_rate = (success / total * 100) if total > 0 else 0
        
        return {
            'total': total,
            'success': success,
            'failed': failed,
            'pending': total - success - failed,
            'success_rate': success_rate,
            'by_platform': by_platform,
            'by_resume': by_resume,
        }
    
    
    def get_recent_applications(self, limit: int = 10) -> List[Dict]:
        """
        Get most recent applications.
        
        Args:
            limit: Number of recent applications to retrieve
            
        Returns:
            List of application dictionaries
        """
        sheet = self.workbook["Applications"]
        
        applications = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        # Get last N rows (most recent)
        for row in rows[-limit:]:
            applications.append({
                'date': row[0],
                'time': row[1],
                'job_title': row[2],
                'company': row[3],
                'platform': row[4],
                'location': row[5],
                'resume_used': row[6],
                'match_score': row[7],
                'confidence': row[8],
                'status': row[9],
                'url': row[10],
                'notes': row[11],
            })
        
        return list(reversed(applications))  # Most recent first
    
    
    def log_error(
        self,
        job_title: str,
        company: str,
        platform: str,
        error_message: str,
        resume_used: str = "N/A",
        match_score: int = 0,
        confidence: float = 0.0
    ):
        """
        Log a failed application attempt.
        """
        self.log_application(
            job_title=job_title,
            company=company,
            platform=platform,
            resume_used=resume_used,
            match_score=match_score,
            confidence=confidence,
            status="Failed",
            notes=f"Error: {error_message}"
        )
    
    
    def close(self):
        """
        Save and close the workbook.
        """
        self.workbook.save(self.excel_path)
        self.workbook.close()
        print("💾 Logger saved and closed")


# ============================================================
# 🧪 TESTING / STANDALONE USAGE
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("🧪 TESTING JOB LOGGER")
    print("=" * 60 + "\n")
    
    # Initialize logger
    logger = JobLogger("data/test_tracker.xlsx")
    
    # Log some test applications
    print("\n📝 Logging test applications...")
    print("-" * 60)
    
    test_applications = [
        {
            'job_title': 'Senior React Developer',
            'company': 'Google',
            'platform': 'LinkedIn',
            'resume_used': 'frontend.pdf',
            'match_score': 92,
            'confidence': 0.89,
            'status': 'Success',
            'location': 'Remote',
            'application_url': 'https://linkedin.com/jobs/123456',
        },
        {
            'job_title': 'Backend Engineer',
            'company': 'Microsoft',
            'platform': 'Indeed',
            'resume_used': 'backend.pdf',
            'match_score': 85,
            'confidence': 0.82,
            'status': 'Success',
            'location': 'Bangalore',
        },
        {
            'job_title': 'Full Stack Developer',
            'company': 'Startup XYZ',
            'platform': 'Wellfound',
            'resume_used': 'fullstack.pdf',
            'match_score': 78,
            'confidence': 0.75,
            'status': 'Failed',
            'notes': 'Form submission timeout',
        },
    ]
    
    for app in test_applications:
        logger.log_application(**app)
    
    # Update daily summary
    print("\n📊 Updating daily summary...")
    print("-" * 60)
    logger.update_daily_summary()
    
    # Get statistics
    print("\n📈 Statistics:")
    print("-" * 60)
    stats = logger.get_statistics()
    print(f"Total Applications: {stats['total']}")
    print(f"Success: {stats['success']}")
    print(f"Failed: {stats['failed']}")
    print(f"Success Rate: {stats['success_rate']:.1f}%")
    
    print("\nBy Platform:")
    for platform, count in stats['by_platform'].items():
        print(f"  {platform}: {count}")
    
    print("\nBy Resume:")
    for resume, count in stats['by_resume'].items():
        print(f"  {resume}: {count}")
    
    # Get today's count
    print("\n📅 Today's Applications:")
    print("-" * 60)
    today_count = logger.get_today_count()
    print(f"Total today: {today_count}")
    
    # Get recent applications
    print("\n🕐 Recent Applications:")
    print("-" * 60)
    recent = logger.get_recent_applications(limit=3)
    for app in recent:
        print(f"  • {app['job_title']} at {app['company']} - {app['status']}")
    
    # Close logger
    logger.close()
    
    print("\n" + "=" * 60)
    print("✅ Testing complete!")
    print(f"📊 Check the Excel file: data/test_tracker.xlsx")
    print("=" * 60)