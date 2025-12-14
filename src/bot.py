"""
🕹️ BOT CONTROLLER
Handles all browser automation using Playwright.
Combines High-Stealth settings with Smart Form Filling logic.
"""

from playwright.sync_api import sync_playwright, BrowserContext, Page
from pathlib import Path
from typing import Optional, Dict
import time

# Import your configuration and utils
from src.config import CHROME_USER_DATA, DRY_RUN
from src.user_config import PROFILE, ANSWERS
from src.utils import (
    human_sleep, 
    human_type, 
    human_scroll, 
    human_click,
    human_mouse_move,
    simulate_reading_pattern
)

class JobBot:
    def __init__(self, headless: bool = False):
        """
        Initialize the bot with browser settings.
        """
        self.headless = headless
        self.profile_dir = Path(CHROME_USER_DATA)
        self.playwright = None
        self.browser: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        # Create profile dir if missing
        self.profile_dir.mkdir(parents=True, exist_ok=True)
        print(f"🤖 Bot initialized (Headless: {headless})")

    def start_browser(self):
        """
        Launch the browser with MAXIMUM anti-detection settings.
        """
        print("🚀 Launching Stealth Browser...")
        self.playwright = sync_playwright().start()
        
        # 1. Launch with specific arguments to hide automation
        self.browser = self.playwright.chromium.launch_persistent_context(
            user_data_dir=str(self.profile_dir),
            channel="chrome",       # Uses your real Chrome
            headless=self.headless,
            viewport={'width': 1920, 'height': 1080},
            args=[
                '--disable-blink-features=AutomationControlled', # CRITICAL: Hides "controlled by automated software"
                '--start-maximized',
                '--disable-infobars',
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--disable-dev-shm-usage',
                '--disable-web-security',
                '--disable-features=IsolateOrigins,site-per-process',
            ]
        )
        
        # 2. Get the page
        self.page = self.browser.pages[0] if self.browser.pages else self.browser.new_page()

        # 3. Inject JavaScript to fake "navigator" properties (The Cloak)
        self.page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => false });
            window.chrome = { runtime: {} };
            human_sleep(0.7, 1.5)
            Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
        """)
        
        print("✅ Stealth Browser Ready")

    def apply_to_job(self, job_url, resume_path):
        """
        The main logic loop to apply for a single job.
        """
        if not self.page:
            self.start_browser()

        print(f"\n🔗 Navigating to: {job_url}")
        try:
            self.page.goto(job_url, timeout=60000)
            human_sleep(1, 2) # Let page load

            # 1. Simulate Reading (Important for stealth)
            simulate_reading_pattern(self.page, "h1")
            # We look for multiple variations of the button
            apply_locators = [
                # Most specific first (single button match)
                "button.jobs-apply-button",
                "button[aria-label*='Easy Apply to']",  # Includes job title - more specific
                "button[data-control-name*='jobdetails_topcard_inapply']",
                # Link variants
                "a[aria-label*='Easy Apply to'][data-view-name='job-apply-button']",
                # Generic fallbacks
                "button[aria-label*='Easy Apply']",
                "a[aria-label*='Easy Apply']",
            ]
            
            clicked = False
            for selector in apply_locators:
                try:
                    loc = self.page.locator(selector)
                    count = loc.count()
                    if count > 0:
                        # Find the first VISIBLE element
                        for i in range(count):
                            elem = loc.nth(i)
                            if elem.is_visible():
                                print(f"👇 Clicking Easy Apply... (found at index {i})")
                                elem.click()
                                clicked = True
                                break
                        if clicked:
                            break
                except Exception as e:
                    continue

            if not clicked:
                print("⚠️ No 'Easy Apply' button found (Might be external or already applied). Skipping.")
                return "Skipped"

            # Wait for the Easy Apply modal to appear
            print("⏳ Waiting for Easy Apply modal...")
            try:
                # Wait for modal container to be visible
                modal_selectors = [
                    ".jobs-easy-apply-content",
                    "div[role='dialog']",
                    ".jobs-easy-apply-modal"
                ]
                modal_found = False
                for selector in modal_selectors:
                    try:
                        self.page.wait_for_selector(selector, state="visible", timeout=5000)
                        print(f"✅ Modal loaded: {selector}")
                        modal_found = True
                        break
                    except:
                        continue
                
                if not modal_found:
                    print("⚠️ Easy Apply modal didn't load. Job might require external application.")
                    return "Skipped (No modal)"
                
                human_sleep(0.5, 1)
            except Exception as e:
                print(f"⚠️ Modal wait failed: {e}")
                return "Skipped (Modal error)"

            # 3. The Form Loop (Handle Popups)
            # We loop up to 10 times to handle multi-page forms (Contact -> Resume -> Review -> Submit)
            max_steps = 10
            for step in range(max_steps):
                print(f"   ➡️ Form Step {step + 1}")
                
                # Scroll within the modal to load all fields
                try:
                    modal = self.page.locator(".jobs-easy-apply-content")
                    if modal.count() > 0:
                        modal.first.evaluate("el => el.scrollBy(0, el.scrollHeight / 2)")
                        human_sleep(0.2, 0.4)
                except:
                    pass
                
                # A. Auto-Fill inputs
                self._fill_smart_fields()
                
                # A2. Handle dropdowns
                self._fill_dropdowns()

                # B. Upload Resume if asked
                self._handle_upload(resume_path)

                # C. Check for SUBMIT (look within modal first)
                modal = self.page.locator(".jobs-easy-apply-content")
                submit_btn = modal.locator("button[aria-label='Submit application']") if modal.count() > 0 else self.page.locator("button[aria-label='Submit application']")
                if submit_btn.count() > 0 and submit_btn.is_visible():
                    print("✅ Found Submit button!")
                    if not DRY_RUN:
                        try:
                            human_click(self.page, "button[aria-label='Submit application']")
                            human_sleep(1, 2) # Wait for submission
                            print("   ✅ Submit clicked")
                        except Exception as e:
                            print("   ❌ Submit click failed:", e)
                            return f"Failed (Submit error: {e})"
                    else:
                        print("   (DRY RUN: Submit skipped)")
                    return "Success"

                # D. Check for NEXT or REVIEW (within modal)
                # Get modal context first
                modal = self.page.locator(".jobs-easy-apply-content")
                search_context = modal if modal.count() > 0 else self.page
                
                # Try multiple Next button selectors
                next_selectors = [
                    "button[aria-label='Continue to next step']",
                    "button[data-easy-apply-next-button]",
                    "button:has-text('Next')",
                    "button[aria-label='Next']",
                ]
                review_selectors = [
                    "button[data-live-test-easy-apply-review-button]",
                    "button[aria-label='Review your application']",
                    "button[aria-label='Review']",
                    "button:has-text('Review')",
                ]

                next_clicked = False
                for selector in next_selectors:
                    try:
                        loc = search_context.locator(selector)
                        if loc.count() > 0 and loc.first.is_visible():
                            try:
                                # Wait for button to be clickable and click it
                                loc.first.scroll_into_view_if_needed()
                                human_sleep(0.2, 0.4)
                                loc.first.click(timeout=5000)
                                print("➡️ Clicked Next")
                                next_clicked = True
                                # Wait for page transition
                                human_sleep(0.8, 1.5)
                                break
                            except Exception as click_err:
                                # If click fails, try next selector
                                continue
                    except Exception:
                        continue

                if next_clicked:
                    pass  # Already clicked, continue to next iteration
                else:
                    # Try Review button
                    review_clicked = False
                    for selector in review_selectors:
                        try:
                            loc = search_context.locator(selector)
                            if loc.count() > 0 and loc.first.is_visible():
                                try:
                                    # Scroll button into view first
                                    loc.first.scroll_into_view_if_needed()
                                    human_sleep(0.2, 0.4)
                                    loc.first.click(timeout=5000)
                                    print("👀 Clicked Review")
                                    review_clicked = True
                                    # Wait for page transition
                                    human_sleep(0.8, 1.5)
                                    break
                                except Exception as click_err:
                                    continue
                        except Exception as e:
                            continue
                    
                    if not review_clicked:
                        # Check for errors
                        if self.page.locator(".artdeco-inline-feedback__message").count() > 0:
                            print("❌ Form Error: Missing required field.")
                            return "Failed (Form Error)"
                        
                        # Debug: show what buttons are visible IN THE MODAL
                        print("⚠️ Stuck: No Next/Submit/Review button found.")
                        print("   🔍 Debugging - checking buttons in modal:")
                        try:
                            modal_buttons = search_context.locator("button").all()
                            for i, btn in enumerate(modal_buttons[:10]):
                                try:
                                    if btn.is_visible():
                                        label = btn.get_attribute("aria-label") or btn.inner_text()[:30] or "no-label"
                                        print(f"      {i+1}. {label[:60]}")
                                except:
                                    pass
                        except:
                            pass
                        return "Failed (Stuck)"
                
                human_sleep(0.3, 0.6)

            return "Failed (Too many steps)"
        except Exception as e:
            print(f"❌ Error applying: {e}")
            return f"Failed: {str(e)}"

    def _fill_smart_fields(self):
        """
        Scans the page for inputs and fills them using USER_CONFIG.
        Tracks unfilled fields for later review.
        """
        unfilled_fields = []
        
        try:
            # 1. Text Inputs
            inputs = self.page.locator("input[type='text'], input[type='email'], input[type='tel'], input[type='number']")
            count = inputs.count()
            
            for i in range(count):
                field = inputs.nth(i)
                if field.is_visible() and not field.get_attribute("value"):
                    label = self._get_label(field).lower()
                    
                    # Match logic - try both directions
                    matched = False
                    for key, value in PROFILE.items():
                        # Check if key is in label OR label contains key words
                        key_lower = key.lower().replace("_", " ")
                        if key_lower in label or any(word in label for word in key_lower.split()):
                            print(f"      ✍️ Filling {key}...")
                            field.fill(str(value))
                            human_sleep(0.2, 0.4)
                            matched = True
                            break
                    
                    if not matched and label:
                        unfilled_fields.append(label)
                        print(f"      ⚠️ Skipped unfilled field: {label}")

            # 2. Radio Buttons (Yes/No)
            fieldsets = self.page.locator("fieldset")
            for i in range(fieldsets.count()):
                group = fieldsets.nth(i)
                text = group.text_content().lower()
                
                matched = False
                for question, answer in ANSWERS.items():
                    if question in text:
                        # Try to click the specific radio (label containing 'Yes' or 'No')
                        option = group.locator(f"label:has-text('{answer}')")
                        if option.is_visible():
                            option.click()
                            human_sleep(0.2, 0.4)
                            matched = True
                            break
                
                if not matched and text.strip():
                    unfilled_fields.append(f"Radio: {text[:100]}")
                    print(f"      ⚠️ Skipped unfilled radio: {text[:100]}")
        
        except Exception as e:
            print(f"⚠️ Fill fields error: {e}")
        
        # Log unfilled fields
        if unfilled_fields:
            self._log_unfilled_fields(unfilled_fields)
    
    def _fill_dropdowns(self):
        """Handle dropdown/select fields."""
        try:
            # Common dropdown answers
            dropdown_answers = {
                "english": "Professional working proficiency",
                "proficiency": "Professional working proficiency",
                "degree": "No",
                "bachelor": "No",
                "master": "No",
                "education": "No",
                "time zone": "No",
                "us time": "No",
                "experience": "2",
                "years": "2",
            }
            
            selects = self.page.locator("select")
            for i in range(selects.count()):
                select = selects.nth(i)
                if select.is_visible():
                    # Get associated label
                    select_id = select.get_attribute("id")
                    label_text = ""
                    if select_id:
                        try:
                            label = self.page.locator(f"label[for='{select_id}']")
                            if label.count() > 0:
                                label_text = label.inner_text().lower()
                        except:
                            pass
                    
                    # Try to match and select
                    for keyword, value in dropdown_answers.items():
                        if keyword in label_text:
                            try:
                                select.select_option(label=value)
                                print(f"      📋 Selected dropdown: {value} for {label_text[:40]}")
                                human_sleep(0.2, 0.4)
                                break
                            except:
                                # Try selecting by value/index if label doesn't work
                                try:
                                    options = select.locator("option").all()
                                    for opt in options:
                                        if value.lower() in opt.inner_text().lower():
                                            opt.click()
                                            print(f"      📋 Selected dropdown: {opt.inner_text()}")
                                            break
                                except:
                                    pass
        except Exception as e:
            print(f"⚠️ Dropdown fill error: {e}")
    
    def _log_unfilled_fields(self, fields: list):
        """Log unfilled fields to a clean, actionable Excel tracker."""
        import pandas as pd
        from pathlib import Path
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        try:
            error_file = Path("data/unfilled_fields_tracker.xlsx")
            
            # Clean up field names for better readability
            cleaned_fields = []
            for field in fields:
                # Remove "Radio: " prefix and clean up text
                clean = field.replace("Radio: ", "").strip()
                # Skip duplicates and empty
                if clean and clean not in cleaned_fields:
                    cleaned_fields.append(clean)
            
            if not cleaned_fields:
                return
            
            # Prepare clean data
            now = datetime.now()
            job_url = self.page.url
            
            # Group by job URL to avoid duplicates
            new_entries = []
            for field in cleaned_fields:
                new_entries.append({
                    "Date": now.strftime("%Y-%m-%d %H:%M"),
                    "Question/Field": field,
                    "Suggested Answer": self._suggest_answer(field),
                    "Job URL": job_url
                })
            
            # Load existing or create new
            if error_file.exists():
                existing_df = pd.read_excel(error_file)
                # Remove duplicates from same job
                existing_df = existing_df[existing_df["Job URL"] != job_url]
                df = pd.concat([existing_df, pd.DataFrame(new_entries)], ignore_index=True)
            else:
                df = pd.DataFrame(new_entries)
            
            # Save to Excel with formatting
            df.to_excel(error_file, index=False)
            
            # Apply formatting
            wb = openpyxl.load_workbook(error_file)
            ws = wb.active
            
            # Header formatting
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            
            for col_num in range(1, 5):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column widths
            ws.column_dimensions['A'].width = 18  # Date
            ws.column_dimensions['B'].width = 60  # Question/Field
            ws.column_dimensions['C'].width = 30  # Suggested Answer
            ws.column_dimensions['D'].width = 50  # Job URL
            
            # Freeze header
            ws.freeze_panes = 'A2'
            
            wb.save(error_file)
            print(f"      📊 Logged {len(cleaned_fields)} unfilled fields to {error_file}")
        except Exception as e:
            print(f"      ⚠️ Could not log unfilled fields: {e}")
    
    def _suggest_answer(self, field: str) -> str:
        """Suggest what to add to user_config based on field name."""
        field_lower = field.lower()
        
        # Salary/CTC questions
        if any(word in field_lower for word in ["ctc", "salary", "compensation", "expected"]):
            if "current" in field_lower:
                return 'Add to PROFILE: "current ctc": "0"'
            else:
                return 'Add to PROFILE: "expected ctc": "18"'
        
        # Experience questions
        elif any(word in field_lower for word in ["experience", "years"]):
            tech = None
            if "python" in field_lower:
                tech = "python"
            elif "java" in field_lower:
                tech = "java"
            elif "react" in field_lower:
                tech = "react"
            elif "node" in field_lower:
                tech = "node"
            elif "aws" in field_lower or "cloud" in field_lower:
                tech = "aws"
            
            if tech:
                return f'Add to PROFILE: "{tech}": "2"'
            else:
                return 'Add to PROFILE: "experience": "2"'
        
        # Notice period
        elif "notice" in field_lower:
            return 'Add to PROFILE: "notice": "0"'
        
        # Yes/No questions (radio/checkboxes)
        elif any(word in field_lower for word in ["willing", "comfortable", "authorized", "completed"]):
            keyword = self._extract_keyword(field_lower)
            return f'Add to ANSWERS: "{keyword}": "Yes" or "No"'
        
        # Dropdown selections
        elif "select an option" in field_lower or "dropdown" in field_lower:
            return "Check dropdown options and add to PROFILE or ANSWERS"
        
        else:
            return "Review question and add appropriate value"
    
    def _extract_keyword(self, text: str) -> str:
        """Extract key phrase from question for ANSWERS dict."""
        # Common patterns
        if "bachelor" in text or "degree" in text:
            return "bachelor"
        elif "background check" in text:
            return "background check"
        elif "remote" in text:
            return "remote"
        elif "relocate" in text:
            return "relocate"
        elif "time zone" in text or "us time" in text:
            return "time zone"
        else:
            # Return first meaningful words
            words = text.split()[:4]
            return " ".join(words)

    def _get_label(self, element):
        """Helper to get label text for an input."""
        try:
            id_val = element.get_attribute("id")
            if id_val:
                return self.page.locator(f"label[for='{id_val}']").inner_text()
        except:
            return ""
        return ""

    def _handle_upload(self, resume_path):
        """Finds file inputs and uploads resume - ALWAYS replaces LinkedIn's stored resume."""
        try:
            file_input = self.page.locator("input[type='file']")
            if file_input.count() > 0:
                # ALWAYS upload our resume, even if LinkedIn has one pre-filled
                print(f"      📎 Uploading resume: {resume_path}")
                file_input.first.set_input_files(resume_path)
                human_sleep(0.5, 1)
                print(f"      ✅ Resume uploaded successfully")
        except Exception as e:
            print(f"      ⚠️ Resume upload skipped: {e}")

    def close(self):
        try:
            # close persistent context / browser
            if getattr(self, 'browser', None):
                try:
                    self.browser.close()
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'playwright', None):
                try:
                    self.playwright.stop()
                except Exception:
                    pass
        except Exception:
            pass

# TEST
if __name__ == "__main__":
    bot = JobBot(headless=False)
    bot.start_browser()
    # bot.apply_to_job("https://linkedin.com/jobs/view/...", "data/resumes/frontend.pdf")